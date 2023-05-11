from openpyxl import load_workbook
from game import *

# Loads a game from an excel file
def getGameFromExcel(filename: str) -> Game:
    wb = load_workbook(filename)
    sheet = wb.active
    event = sheet['B1'].value
    site = sheet['B2'].value
    date = sheet['B3'].value
    rnd = sheet['B4'].value
    white = sheet['B5'].value
    black = sheet['B6'].value
    result = sheet['B7'].value
    eco = sheet['B8'].value
    opening = sheet['B9'].value
    plyCount = sheet['B10'].value
    whiteElo = sheet['B11'].value
    blackElo = sheet['B12'].value
    moves = sheet['B13'].value
    return Game(event, site, date, rnd, white, black, result, eco, opening, plyCount, whiteElo, blackElo, moves)

# Writes a game to an excel file
def writeGameToExcel(filename: str, game: Game):
    wb = load_workbook(filename)
    sheet = wb.active
    #Resets the spreadsheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
    sheet.cell(row=1, column = 1).value= "Event"
    sheet.cell(row=2, column = 1).value= "Site"
    sheet.cell(row=3, column = 1).value= "Date"
    sheet.cell(row=4, column = 1).value= "Rnd"
    sheet.cell(row=5, column = 1).value= "White"
    sheet.cell(row=6, column = 1).value= "Black"
    sheet.cell(row=7, column = 1).value= "Result"
    sheet.cell(row=8, column = 1).value= "Eco"
    sheet.cell(row=9, column = 1).value= "Opening"
    sheet.cell(row=10, column = 1).value= "PlyCount"
    sheet.cell(row=11, column = 1).value= "WhiteElo"
    sheet.cell(row=12, column = 1).value= "BlackElo"
    sheet.cell(row=13, column = 1).value= "Moves"
    sheet.cell(row=1, column = 2).value= game.event
    sheet.cell(row=2, column = 2).value= game.site
    sheet.cell(row=3, column = 2).value= game.date
    sheet.cell(row=4, column = 2).value= game.rnd
    sheet.cell(row=5, column = 2).value= game.white
    sheet.cell(row=6, column = 2).value= game.black
    sheet.cell(row=7, column = 2).value= game.result
    sheet.cell(row=8, column = 2).value= game.eco
    sheet.cell(row=9, column = 2).value= game.opening
    sheet.cell(row=10, column = 2).value= game.plyCount
    sheet.cell(row=11, column = 2).value= game.whiteElo
    sheet.cell(row=12, column = 2).value= game.blackElo
    for i in range(int(game.plyCount)//2):
        rowNew: int = 13 + int(i)
        sheet.cell(row=rowNew, column = 2).value = game.getDuoMovesAsList()[i]
    wb.save(filename)
